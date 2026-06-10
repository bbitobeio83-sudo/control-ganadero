"""Generación de reportes en PDF y Excel para el sistema ganadero."""
import io
import os
from datetime import date

_BASE = os.path.dirname(os.path.abspath(__file__))

# ── PDF ───────────────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer, HRFlowable, PageBreak,
                                Image as RLImage)

# ── Excel ─────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colores PDF ───────────────────────────────────────────────────────────────
C_VERDE      = colors.HexColor('#1a5c26')
C_VERDE_FILA = colors.HexColor('#e8f5e9')
C_GRIS       = colors.HexColor('#f5f5f5')
C_ROJO_FILA  = colors.HexColor('#ffebee')
C_AZUL       = colors.HexColor('#1565c0')


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS PDF
# ══════════════════════════════════════════════════════════════════════════════

def _styles():
    s = getSampleStyleSheet()
    s.add(ParagraphStyle('Rancho',   parent=s['Title'],   textColor=C_VERDE,  fontSize=16))
    s.add(ParagraphStyle('Titulo',   parent=s['Heading2'], textColor=C_VERDE,  spaceAfter=2))
    s.add(ParagraphStyle('Sub',      parent=s['Normal'],   textColor=colors.grey, fontSize=9))
    s.add(ParagraphStyle('SecHead',  parent=s['Heading3'], textColor=C_AZUL,   spaceBefore=8))
    return s


def _pdf_header(elements, titulo, subtitulo='', cfg=None):
    cfg = cfg or {}
    s = _styles()
    nombre      = cfg.get('nombre_hacienda') or 'Mi Rancho'
    propietario = cfg.get('propietario', '')
    ubicacion   = cfg.get('ubicacion', '')
    logo_archivo = cfg.get('logo_archivo', '')

    # ── Logo + nombre en fila horizontal (tabla 2 columnas)
    nombre_parr = [Paragraph(nombre.upper(), s['Rancho'])]
    if propietario or ubicacion:
        linea = ' | '.join(filter(None, [propietario, ubicacion]))
        nombre_parr.append(Paragraph(linea, s['Sub']))

    logo_cell = []
    if logo_archivo:
        logo_path = os.path.join(_BASE, 'static', 'uploads', logo_archivo)
        if os.path.exists(logo_path):
            try:
                logo_img = RLImage(logo_path, width=2.2*cm, height=2.2*cm)
                logo_cell = [logo_img]
            except Exception:
                pass

    if logo_cell:
        head_table = Table(
            [[logo_cell[0], nombre_parr]],
            colWidths=[2.6*cm, None],
            hAlign='LEFT',
        )
        head_table.setStyle(TableStyle([
            ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING',(0,0), (-1,-1), 6),
            ('TOPPADDING',  (0,0), (-1,-1), 0),
            ('BOTTOMPADDING',(0,0),(-1,-1), 0),
        ]))
        elements.append(head_table)
    else:
        for p in nombre_parr:
            elements.append(p)

    elements.append(Paragraph(titulo, s['Titulo']))
    if subtitulo:
        elements.append(Paragraph(subtitulo, s['Sub']))
    elements.append(Paragraph(f'Fecha de generación: {date.today().strftime("%d/%m/%Y")}', s['Sub']))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(HRFlowable(width='100%', color=C_VERDE, thickness=1.5))
    elements.append(Spacer(1, 0.4 * cm))


def _base_style(header_bg=None):
    bg = header_bg or C_VERDE
    return TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), bg),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 8),
        ('FONTSIZE',      (0, 1), (-1, -1), 7.5),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, C_GRIS]),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.lightgrey),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
    ])


def _build_pdf(elements, pagesize=A4):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagesize,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    doc.build(elements)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
#  PDF — ANIMALES
# ══════════════════════════════════════════════════════════════════════════════

def pdf_animales(animales, cfg=None):
    el = []
    _pdf_header(el, 'Inventario de Animales', f'Total: {len(animales)} animal(es)', cfg)

    data = [['Arete', 'Nombre', 'Raza', 'Sexo', 'F. Nacimiento',
             'Peso kg', 'Propósito', 'Estado']]
    for a in animales:
        data.append([
            a.get('arete', ''),
            a.get('nombre') or '—',
            a.get('raza', ''),
            (a.get('sexo') or '').title(),
            a.get('fecha_nacimiento') or '—',
            f"{a['peso_actual']:.1f}" if a.get('peso_actual') else '—',
            (a.get('proposito') or '').replace('_', ' ').title(),
            (a.get('estado') or '').title(),
        ])

    t = Table(data, repeatRows=1, hAlign='LEFT')
    t.setStyle(_base_style())
    el.append(t)
    return _build_pdf(el, landscape(A4))


# ══════════════════════════════════════════════════════════════════════════════
#  PDF — SANIDAD
# ══════════════════════════════════════════════════════════════════════════════

def pdf_salud(registros, cfg=None):
    el = []
    _pdf_header(el, 'Registro Sanitario', f'{len(registros)} registro(s)', cfg)

    data = [['Fecha', 'Arete', 'Nombre', 'Tipo', 'Descripción',
             'Medicamento', 'Dosis', 'Veterinario', 'Costo', 'Próx. Cita']]
    for r in registros:
        data.append([
            r.get('fecha', ''),
            r.get('arete', ''),
            r.get('nombre') or '—',
            (r.get('tipo') or '').title(),
            (r.get('descripcion') or '—')[:28],
            r.get('medicamento') or '—',
            r.get('dosis') or '—',
            r.get('veterinario') or '—',
            f"${r.get('costo', 0):.2f}",
            r.get('proxima_cita') or '—',
        ])

    t = Table(data, repeatRows=1, hAlign='LEFT')
    t.setStyle(_base_style())
    el.append(t)
    return _build_pdf(el, landscape(A4))


# ══════════════════════════════════════════════════════════════════════════════
#  PDF — PRODUCCIÓN DE LECHE
# ══════════════════════════════════════════════════════════════════════════════

def pdf_produccion_dia(registros, fecha, total, cfg=None):
    el = []
    _pdf_header(el, f'Producción de Leche — {fecha}',
                f'Total del día: {total:.1f} litros  |  Vacas ordeñadas: {len(registros)}', cfg)

    data = [['Arete', 'Nombre', 'Mañana (L)', 'Tarde (L)', 'Total (L)']]
    for r in registros:
        data.append([
            r.get('arete', ''), r.get('nombre') or '—',
            f"{r.get('manana', 0):.1f}",
            f"{r.get('tarde', 0):.1f}",
            f"{r.get('total', 0):.1f}",
        ])
    # Fila total
    data.append(['', 'TOTAL DEL DÍA', '', '', f'{total:.1f}'])

    style = _base_style()
    style.add('FONTNAME',   (0, len(data)-1), (-1, len(data)-1), 'Helvetica-Bold')
    style.add('BACKGROUND', (0, len(data)-1), (-1, len(data)-1), C_VERDE_FILA)
    t = Table(data, repeatRows=1, hAlign='LEFT')
    t.setStyle(style)
    el.append(t)
    return _build_pdf(el)


def pdf_produccion_mensual(produccion_mensual, cfg=None):
    el = []
    _pdf_header(el, 'Producción de Leche — Últimos 30 días', cfg=cfg)

    data = [['Fecha', 'Litros Totales', 'Vacas Ordeñadas', 'Promedio por Vaca']]
    for r in produccion_mensual:
        vacas = r.get('vacas', 1) or 1
        litros = r.get('litros', 0) or 0
        data.append([
            r.get('fecha', ''),
            f'{litros:.1f}',
            str(vacas),
            f'{litros / vacas:.1f}',
        ])

    t = Table(data, repeatRows=1, hAlign='LEFT')
    t.setStyle(_base_style())
    el.append(t)
    return _build_pdf(el)


# ══════════════════════════════════════════════════════════════════════════════
#  PDF — FINANZAS
# ══════════════════════════════════════════════════════════════════════════════

def pdf_finanzas(registros, mes, ingresos, gastos, cfg=None):
    el = []
    _pdf_header(el, f'Control Financiero — {mes}',
                f'Ingresos: ${ingresos:,.2f}   |   Gastos: ${gastos:,.2f}'
                f'   |   Balance: ${ingresos - gastos:,.2f}', cfg)

    data = [['Fecha', 'Tipo', 'Categoría', 'Descripción', 'Animal', 'Monto']]
    for r in registros:
        data.append([
            r.get('fecha', ''),
            (r.get('tipo') or '').title(),
            r.get('categoria', ''),
            (r.get('descripcion') or '—')[:32],
            r.get('arete') or '—',
            f"${r.get('monto', 0):,.2f}",
        ])

    style = _base_style()
    for i, r in enumerate(registros, 1):
        bg = C_VERDE_FILA if r.get('tipo') == 'ingreso' else C_ROJO_FILA
        style.add('BACKGROUND', (0, i), (-1, i), bg)

    t = Table(data, repeatRows=1, hAlign='LEFT')
    t.setStyle(style)
    el.append(t)
    return _build_pdf(el, landscape(A4))


# ══════════════════════════════════════════════════════════════════════════════
#  PDF — REPORTE GENERAL
# ══════════════════════════════════════════════════════════════════════════════

def pdf_reporte_general(inv_raza, leche_mensual, balance_mensual,
                        top_productoras, partos_anno, cfg=None):
    el = []
    s = _styles()
    _pdf_header(el, 'Reporte General del Rancho', cfg=cfg)

    # ── Inventario por raza
    el.append(Paragraph('1. Inventario por Raza y Sexo', s['SecHead']))
    el.append(Spacer(1, 0.2*cm))
    data = [['Raza', 'Sexo', 'Cantidad']]
    for r in inv_raza:
        data.append([r.get('raza', ''), (r.get('sexo') or '').title(),
                     str(r.get('cantidad', 0))])
    t = Table(data, repeatRows=1, hAlign='LEFT')
    t.setStyle(_base_style())
    el.append(t)
    el.append(Spacer(1, 0.5*cm))

    # ── Producción leche mensual
    el.append(Paragraph('2. Producción de Leche Mensual', s['SecHead']))
    el.append(Spacer(1, 0.2*cm))
    data = [['Mes', 'Total Litros', 'Promedio Diario']]
    for r in leche_mensual:
        data.append([r.get('mes', ''), f"{r.get('litros', 0):.1f}",
                     f"{r.get('promedio_dia', 0):.1f}"])
    t = Table(data, repeatRows=1, hAlign='LEFT')
    t.setStyle(_base_style())
    el.append(t)
    el.append(Spacer(1, 0.5*cm))

    # ── Balance mensual
    el.append(Paragraph('3. Balance Financiero Mensual', s['SecHead']))
    el.append(Spacer(1, 0.2*cm))
    data = [['Mes', 'Ingresos', 'Gastos', 'Balance']]
    for r in balance_mensual:
        ing = r.get('ingresos', 0)
        gas = r.get('gastos', 0)
        data.append([r.get('mes', ''), f'${ing:,.2f}', f'${gas:,.2f}',
                     f'${ing - gas:,.2f}'])
    style = _base_style()
    for i, r in enumerate(balance_mensual, 1):
        bal = r.get('ingresos', 0) - r.get('gastos', 0)
        style.add('TEXTCOLOR', (3, i), (3, i),
                  C_VERDE if bal >= 0 else colors.red)
    t = Table(data, repeatRows=1, hAlign='LEFT')
    t.setStyle(style)
    el.append(t)
    el.append(Spacer(1, 0.5*cm))

    # ── Top productoras
    el.append(Paragraph('4. Top Productoras — Últimos 30 días', s['SecHead']))
    el.append(Spacer(1, 0.2*cm))
    data = [['#', 'Arete', 'Nombre', 'Raza', 'Total L', 'Prom./Día', 'Días']]
    for i, r in enumerate(top_productoras, 1):
        data.append([str(i), r.get('arete', ''), r.get('nombre') or '—',
                     r.get('raza', ''),
                     f"{r.get('total_litros', 0):.1f}",
                     f"{r.get('promedio_dia', 0):.1f}",
                     str(r.get('dias_ordenada', 0))])
    t = Table(data, repeatRows=1, hAlign='LEFT')
    t.setStyle(_base_style())
    el.append(t)
    el.append(Spacer(1, 0.5*cm))

    # ── Partos por mes
    if partos_anno:
        el.append(Paragraph('5. Partos por Mes (últimos 12 meses)', s['SecHead']))
        el.append(Spacer(1, 0.2*cm))
        data = [['Mes', 'Número de Partos']]
        for r in partos_anno:
            data.append([r.get('mes', ''), str(r.get('partos', 0))])
        t = Table(data, repeatRows=1, hAlign='LEFT')
        t.setStyle(_base_style())
        el.append(t)

    return _build_pdf(el)


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS EXCEL
# ══════════════════════════════════════════════════════════════════════════════

_FILL_VERDE   = PatternFill('solid', fgColor='1a5c26')
_FILL_ALT1    = PatternFill('solid', fgColor='f1f8e9')
_FILL_ALT2    = PatternFill('solid', fgColor='ffffff')
_FILL_ING     = PatternFill('solid', fgColor='e8f5e9')
_FILL_GAS     = PatternFill('solid', fgColor='ffebee')
_FILL_TOTAL   = PatternFill('solid', fgColor='c8e6c9')
_FONT_HDR     = Font(color='FFFFFF', bold=True, size=10)
_FONT_TITLE   = Font(bold=True, size=13, color='1a5c26')
_FONT_SUB     = Font(italic=True, size=9, color='555555')
_FONT_TOTAL   = Font(bold=True, size=10)
_BORDER_THIN  = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def _xl_title(ws, titulo, subtitulo=''):
    ws['A1'] = titulo
    ws['A1'].font = _FONT_TITLE
    ws.row_dimensions[1].height = 24
    if subtitulo:
        ws['A2'] = subtitulo
        ws['A2'].font = _FONT_SUB
        ws.row_dimensions[2].height = 14
    ws['A3'] = f'Generado el {date.today().strftime("%d/%m/%Y")}'
    ws['A3'].font = _FONT_SUB
    ws.append([])  # fila vacía → header en fila 5


def _xl_header(ws, cols):
    ws.append(cols)
    for cell in ws[ws.max_row]:
        cell.fill   = _FILL_VERDE
        cell.font   = _FONT_HDR
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _BORDER_THIN
    ws.row_dimensions[ws.max_row].height = 22


def _xl_autowidth(ws, extra=3):
    for col in ws.columns:
        try:
            max_len = max(len(str(c.value or '')) for c in col if c.value is not None)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + extra, 45)
        except Exception:
            pass


def _xl_row(ws, values, fill_even=None, fill_odd=None, row_num=0):
    ws.append(values)
    row = ws[ws.max_row]
    fill = (fill_even if row_num % 2 == 0 else fill_odd) or _FILL_ALT2
    for cell in row:
        cell.fill   = fill
        cell.border = _BORDER_THIN
        cell.alignment = Alignment(vertical='center')


def _xl_save(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — ANIMALES
# ══════════════════════════════════════════════════════════════════════════════

def excel_animales(animales, cfg=None):
    cfg = cfg or {}
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    ws.sheet_view.showGridLines = False

    nombre = cfg.get('nombre_hacienda') or 'Mi Rancho'
    _xl_title(ws, f'{nombre} — Inventario de Animales',
              f'{len(animales)} animal(es) registrado(s)')
    _xl_header(ws, ['Arete', 'Nombre', 'Raza', 'Sexo', 'F. Nacimiento',
                    'Peso (kg)', 'Propósito', 'Estado', 'F. Ingreso', 'Observaciones'])

    for i, a in enumerate(animales):
        _xl_row(ws, [
            a.get('arete', ''), a.get('nombre') or '',
            a.get('raza', ''), (a.get('sexo') or '').title(),
            a.get('fecha_nacimiento') or '',
            round(a['peso_actual'], 1) if a.get('peso_actual') else '',
            (a.get('proposito') or '').replace('_', ' ').title(),
            (a.get('estado') or '').title(),
            a.get('fecha_ingreso') or '',
            a.get('observaciones') or '',
        ], _FILL_ALT1, _FILL_ALT2, i)

    _xl_autowidth(ws)
    return _xl_save(wb)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — SANIDAD
# ══════════════════════════════════════════════════════════════════════════════

def excel_salud(registros, cfg=None):
    cfg = cfg or {}
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sanidad'
    ws.sheet_view.showGridLines = False

    nombre = cfg.get('nombre_hacienda') or 'Mi Rancho'
    _xl_title(ws, f'{nombre} — Registro Sanitario', f'{len(registros)} registro(s)')
    _xl_header(ws, ['Fecha', 'Arete', 'Nombre', 'Tipo', 'Descripción',
                    'Medicamento', 'Dosis', 'Vía', 'Veterinario', 'Costo ($)', 'Próx. Cita'])

    alt = PatternFill('solid', fgColor='f3e5f5')
    for i, r in enumerate(registros):
        _xl_row(ws, [
            r.get('fecha', ''), r.get('arete', ''), r.get('nombre') or '',
            (r.get('tipo') or '').title(), r.get('descripcion') or '',
            r.get('medicamento') or '', r.get('dosis') or '',
            r.get('via_administracion') or '', r.get('veterinario') or '',
            round(r.get('costo', 0), 2), r.get('proxima_cita') or '',
        ], _FILL_ALT1, alt, i)

    _xl_autowidth(ws)
    return _xl_save(wb)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — PRODUCCIÓN
# ══════════════════════════════════════════════════════════════════════════════

def excel_produccion(registros_dia, produccion_mensual, top_vacas, fecha, total, cfg=None):
    cfg = cfg or {}
    wb = Workbook()
    nombre = cfg.get('nombre_hacienda') or 'Mi Rancho'

    # Hoja 1: día seleccionado
    ws1 = wb.active
    ws1.title = f'Día {fecha}'
    ws1.sheet_view.showGridLines = False
    _xl_title(ws1, f'{nombre} — Producción de Leche — {fecha}',
              f'Total: {total:.1f} litros  |  Vacas: {len(registros_dia)}')
    _xl_header(ws1, ['Arete', 'Nombre', 'Mañana (L)', 'Tarde (L)', 'Total (L)'])
    for i, r in enumerate(registros_dia):
        _xl_row(ws1, [r.get('arete', ''), r.get('nombre') or '',
                      r.get('manana', 0), r.get('tarde', 0), r.get('total', 0)],
                _FILL_ALT1, _FILL_ALT2, i)
    # Fila total
    ws1.append(['', 'TOTAL', '', '', round(total, 1)])
    for cell in ws1[ws1.max_row]:
        cell.fill   = _FILL_TOTAL
        cell.font   = _FONT_TOTAL
        cell.border = _BORDER_THIN
    _xl_autowidth(ws1)

    # Hoja 2: últimos 30 días
    ws2 = wb.create_sheet('Últimos 30 días')
    ws2.sheet_view.showGridLines = False
    _xl_title(ws2, 'Producción — Últimos 30 días')
    _xl_header(ws2, ['Fecha', 'Litros Totales', 'Vacas Ordeñadas', 'Prom. por Vaca'])
    for i, r in enumerate(produccion_mensual):
        vacas  = r.get('vacas', 1) or 1
        litros = r.get('litros', 0) or 0
        _xl_row(ws2, [r.get('fecha', ''), round(litros, 1), vacas,
                      round(litros / vacas, 1)], _FILL_ALT1, _FILL_ALT2, i)
    _xl_autowidth(ws2)

    # Hoja 3: top productoras
    ws3 = wb.create_sheet('Top Productoras')
    ws3.sheet_view.showGridLines = False
    _xl_title(ws3, 'Top Productoras — Últimos 30 días')
    _xl_header(ws3, ['#', 'Arete', 'Nombre', 'Total (L)'])
    for i, r in enumerate(top_vacas):
        _xl_row(ws3, [i + 1, r.get('arete', ''), r.get('nombre') or '',
                      round(r.get('total_litros', 0), 1)], _FILL_ALT1, _FILL_ALT2, i)
    _xl_autowidth(ws3)

    return _xl_save(wb)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — FINANZAS
# ══════════════════════════════════════════════════════════════════════════════

def excel_finanzas(registros, mes, ingresos, gastos, cfg=None):
    cfg = cfg or {}
    wb = Workbook()
    ws = wb.active
    ws.title = 'Movimientos'
    ws.sheet_view.showGridLines = False

    nombre = cfg.get('nombre_hacienda') or 'Mi Rancho'
    _xl_title(ws, f'{nombre} — Control Financiero — {mes}',
              f'Ingresos: ${ingresos:,.2f}  |  Gastos: ${gastos:,.2f}'
              f'  |  Balance: ${ingresos - gastos:,.2f}')
    _xl_header(ws, ['Fecha', 'Tipo', 'Categoría', 'Descripción', 'Animal', 'Monto ($)'])

    for r in registros:
        ws.append([
            r.get('fecha', ''),
            (r.get('tipo') or '').title(),
            r.get('categoria', ''),
            r.get('descripcion') or '',
            r.get('arete') or '',
            round(r.get('monto', 0), 2),
        ])
        fill = _FILL_ING if r.get('tipo') == 'ingreso' else _FILL_GAS
        for cell in ws[ws.max_row]:
            cell.fill   = fill
            cell.border = _BORDER_THIN
            cell.alignment = Alignment(vertical='center')

    _xl_autowidth(ws)
    return _xl_save(wb)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — REPORTE GENERAL (múltiples hojas)
# ══════════════════════════════════════════════════════════════════════════════

def excel_reporte_general(inv_raza, leche_mensual, balance_mensual,
                          top_productoras, partos_anno, cfg=None):
    cfg = cfg or {}
    wb = Workbook()
    nombre = cfg.get('nombre_hacienda') or 'Mi Rancho'

    # Hoja 1: Inventario por raza
    ws1 = wb.active
    ws1.title = 'Inventario Raza'
    ws1.sheet_view.showGridLines = False
    _xl_title(ws1, f'{nombre} — Inventario por Raza y Sexo')
    _xl_header(ws1, ['Raza', 'Sexo', 'Cantidad'])
    for i, r in enumerate(inv_raza):
        _xl_row(ws1, [r.get('raza', ''), (r.get('sexo') or '').title(),
                      r.get('cantidad', 0)], _FILL_ALT1, _FILL_ALT2, i)
    _xl_autowidth(ws1)

    # Hoja 2: Producción mensual
    ws2 = wb.create_sheet('Producción Mensual')
    ws2.sheet_view.showGridLines = False
    _xl_title(ws2, 'Producción de Leche Mensual')
    _xl_header(ws2, ['Mes', 'Total Litros', 'Promedio Diario'])
    for i, r in enumerate(leche_mensual):
        _xl_row(ws2, [r.get('mes', ''), round(r.get('litros', 0), 1),
                      round(r.get('promedio_dia', 0), 1)], _FILL_ALT1, _FILL_ALT2, i)
    _xl_autowidth(ws2)

    # Hoja 3: Balance mensual
    ws3 = wb.create_sheet('Balance Mensual')
    ws3.sheet_view.showGridLines = False
    _xl_title(ws3, 'Balance Financiero Mensual (12 meses)')
    _xl_header(ws3, ['Mes', 'Ingresos ($)', 'Gastos ($)', 'Balance ($)'])
    for i, r in enumerate(balance_mensual):
        ing = r.get('ingresos', 0)
        gas = r.get('gastos', 0)
        ws3.append([r.get('mes', ''), round(ing, 2), round(gas, 2), round(ing - gas, 2)])
        row = ws3[ws3.max_row]
        fill = _FILL_ING if ing - gas >= 0 else _FILL_GAS
        for cell in row:
            cell.fill   = fill
            cell.border = _BORDER_THIN
            cell.alignment = Alignment(vertical='center')
    _xl_autowidth(ws3)

    # Hoja 4: Top productoras
    ws4 = wb.create_sheet('Top Productoras')
    ws4.sheet_view.showGridLines = False
    _xl_title(ws4, 'Top Productoras — Últimos 30 días')
    _xl_header(ws4, ['#', 'Arete', 'Nombre', 'Raza', 'Total L', 'Prom./Día', 'Días Ordeñada'])
    for i, r in enumerate(top_productoras):
        _xl_row(ws4, [
            i + 1, r.get('arete', ''), r.get('nombre') or '', r.get('raza', ''),
            round(r.get('total_litros', 0), 1),
            round(r.get('promedio_dia', 0), 1),
            r.get('dias_ordenada', 0),
        ], _FILL_ALT1, _FILL_ALT2, i)
    _xl_autowidth(ws4)

    # Hoja 5: Partos
    ws5 = wb.create_sheet('Partos por Mes')
    ws5.sheet_view.showGridLines = False
    _xl_title(ws5, 'Partos por Mes (últimos 12 meses)')
    _xl_header(ws5, ['Mes', 'Número de Partos'])
    for i, r in enumerate(partos_anno):
        _xl_row(ws5, [r.get('mes', ''), r.get('partos', 0)], _FILL_ALT1, _FILL_ALT2, i)
    _xl_autowidth(ws5)

    return _xl_save(wb)
